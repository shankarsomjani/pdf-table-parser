import os
import time
import streamlit as st
import pandas as pd
import io
import zipfile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import pdfplumber

from adobe.pdfservices.operation.auth.service_principal_credentials import ServicePrincipalCredentials
from adobe.pdfservices.operation.exception.exceptions import ServiceApiException, ServiceUsageException, SdkException
from adobe.pdfservices.operation.io.stream_asset import StreamAsset
from adobe.pdfservices.operation.pdf_services import PDFServices
from adobe.pdfservices.operation.pdf_services_media_type import PDFServicesMediaType
from adobe.pdfservices.operation.pdfjobs.jobs.extract_pdf_job import ExtractPDFJob
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_element_type import ExtractElementType
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_pdf_params import ExtractPDFParams
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_renditions_element_type import ExtractRenditionsElementType
from adobe.pdfservices.operation.pdfjobs.result.extract_pdf_result import ExtractPDFResult

# Import OCR parameters
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.ocr_params import OcrParams
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.ocr_language import OcrLanguage

# --- Function to replace _x000D_ and other unwanted characters ---
def replace_x000d(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cleaned_value = cell.value.replace('_x000D_', '').replace('\r', ' ').replace('\n', ' ').strip()
                cell.value = cleaned_value
    
    return wb

# --- Function to apply company-specific mappings ---
def apply_company_mappings(df, company, mapping_df):
    if df.empty or df.columns.empty:
        return df
    
    company_map = mapping_df[mapping_df['Company'].str.lower() == company.lower()]
    
    if company_map.empty:
        return df
    
    replace_dict = {}
    for _, row in company_map.iterrows():
        original = row['Original']
        mapped = row['Mapped']
        if original and isinstance(original, str) and mapped:
            replace_dict[original.lower()] = mapped
    
    df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: replace_dict.get(str(x).lower(), x) if pd.notna(x) else x)
    return df

# --- Adobe Table Formatter ---
def merge_adobe_tables(zip_path: str) -> bytes:
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Tables"

    table_count = 1
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        file_names = sorted([f for f in zip_ref.namelist() if f.endswith(".xlsx")])
        for file in file_names:
            with zip_ref.open(file) as f:
                df = pd.read_excel(f)
                if df.empty:
                    continue
                ws.append([f"Table {table_count}"])
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                ws.append([])
                ws.append([])
                table_count += 1

    wb.save(output)
    output.seek(0)
    return output.getvalue()

# --- Function to Extract PDF using Adobe PDF Services with optional OCR ---
def extract_pdf_with_adobe(uploaded_pdf, use_ocr=False):
    credentials = ServicePrincipalCredentials(
        client_id=os.getenv("PDF_SERVICES_CLIENT_ID"),
        client_secret=os.getenv("PDF_SERVICES_CLIENT_SECRET")
    )
    pdf_services = PDFServices(credentials=credentials)
    input_asset = pdf_services.upload(input_stream=uploaded_pdf, mime_type=PDFServicesMediaType.PDF)

    if use_ocr:
        ocr_params = OcrParams(
            ocr=True,
            ocr_language=OcrLanguage.ENGLISH  # Change or make dynamic if needed
        )
        extract_pdf_params = ExtractPDFParams(
            elements_to_extract=[ExtractElementType.TEXT, ExtractElementType.TABLES],
            elements_to_extract_renditions=[ExtractRenditionsElementType.TABLES],
            add_char_info=True,
            ocr_params=ocr_params
        )
    else:
        extract_pdf_params = ExtractPDFParams(
            elements_to_extract=[ExtractElementType.TEXT, ExtractElementType.TABLES],
            elements_to_extract_renditions=[ExtractRenditionsElementType.TABLES],
            add_char_info=True,
        )

    extract_pdf_job = ExtractPDFJob(input_asset=input_asset, extract_pdf_params=extract_pdf_params)
    location = pdf_services.submit(extract_pdf_job)

    try:
        pdf_services_response = pdf_services.get_job_result(location, ExtractPDFResult)
        if pdf_services_response is None or pdf_services_response.get_result() is None:
            raise ValueError("Adobe PDF Services did not return valid results.")

        result_asset = pdf_services_response.get_result().get_resource()
        stream_asset: StreamAsset = pdf_services.get_content(result_asset)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        zip_path = f"/tmp/output_adobe_{timestamp}.zip"
        with open(zip_path, "wb") as out_file:
            out_file.write(stream_asset.get_input_stream())

        return zip_path

    except Exception as e:
        raise ValueError(f"Error occurred while extracting PDF using Adobe API: {str(e)}")

# --- Function to process Standard extraction ---
def extract_standard_pdf(uploaded_pdf):
    with pdfplumber.open(uploaded_pdf) as pdf:
        all_tables = []
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for idx, table in enumerate(tables):
                if table:
                    df = pd.DataFrame(table[1:], columns=table[0]) if len(table) > 1 else pd.DataFrame(table)
                    all_tables.append((f"Page{page_num}_Table{idx+1}", df))
    return all_tables

# --- Function to process LLM extraction ---
def extract_llm_pdf(uploaded_pdf, api_key):
    whisperer = LLMWhispererClientV2(api_key=api_key, logging_level="DEBUG")
    temp_path = "/tmp/uploaded_llm.pdf"
    with open(temp_path, "wb") as f:
        f.write(uploaded_pdf.read())

    job_info = whisperer.whisper(
        file_path=temp_path,
        filename=uploaded_pdf.name,
        mode="form",
        output_mode="layout_preserving"
    )

    whisper_hash = job_info.get("whisper_hash")
    if not whisper_hash:
        raise ValueError("Failed to initiate LLMWhisperer job.")

    st.info("⏳ Waiting for LLMWhisperer to process the file...")
    status = None
    for _ in range(20):
        status_info = whisperer.whisper_status(whisper_hash=whisper_hash)
        status = status_info.get("status")
        if status == "processed":
            break
        elif status == "error":
            raise ValueError("LLMWhisperer reported an error while processing the document.")
        time.sleep(2)

    if status != "processed":
        raise ValueError("Timed out waiting for LLMWhisperer to finish processing.")

    result = whisperer.whisper_retrieve(whisper_hash=whisper_hash)
    return result

# --- Streamlit Page Setup ---
st.set_page_config(page_title="PDF Table Extractor & Excel Updater", layout="centered")
st.title("\U0001F4C4 PDF Table Extractor & Excel Updater")

# --- Company Mapping CSV File ---
mapping_df = pd.read_csv("company_mappings.csv") if os.path.exists("company_mappings.csv") else pd.DataFrame(columns=['Company', 'Original', 'Mapped'])
companies = sorted(mapping_df['Company'].unique()) if not mapping_df.empty else []

selected_company = st.selectbox("Select the company:", companies) if companies else None

# --- Upload PDF file ---
uploaded_pdf = st.file_uploader("Upload a PDF file", type="pdf")

# --- Mode selection ---
mode = st.radio("Choose extraction mode:", ["Standard (Code-based)", "LLM (via LLMWhisperer)", "Adobe PDF Services"])

if uploaded_pdf:
    if mode == "Adobe PDF Services":
        use_ocr = st.checkbox("Enable OCR (for scanned PDFs)", value=False)
        try:
            st.info("⏳ Extracting using Adobe PDF Services...")
            zip_path = extract_pdf_with_adobe(uploaded_pdf, use_ocr=use_ocr)

            excel_bytes = merge_adobe_tables(zip_path)

            wb = replace_x000d(io.BytesIO(excel_bytes))
            sheet = wb.active
            data = sheet.values
            columns = next(data)[0:]
            df = pd.DataFrame(data, columns=columns)

            df = df.fillna('')

            if selected_company and not mapping_df.empty:
                df = apply_company_mappings(df, selected_company, mapping_df)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet.title)

            output.seek(0)
            st.download_button("📊 Download Cleaned and Updated Excel", output, "cleaned_and_updated.xlsx")

            st.success("✅ Excel file cleaned and updated successfully!")

        except Exception as e:
            st.error(f"❌ Error processing the file: {str(e)}")

    elif mode == "Standard (Code-based)":
        try:
            st.info("⏳ Extracting using Standard method...")
            all_tables = extract_standard_pdf(uploaded_pdf)

            if all_tables:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for name, df in all_tables:
                        df.to_excel(writer, sheet_name=name[:31], index=False)
                st.success(f"✅ Extracted {len(all_tables)} table(s) using Standard method")
                st.download_button("📅 Download Standard Extracted Excel", output.getvalue(), "standard_tables.xlsx")
            else:
                st.warning("⚠️ No tables found using standard method.")

        except Exception as e:
            st.error(f"❌ Error processing Standard extraction: {str(e)}")

    elif mode == "LLM (via LLMWhisperer)":
        try:
            st.info("⏳ Extracting using LLMWhisperer...")
            LLM_API_KEY = st.secrets.get("LLM_API_KEY")
            if not LLM_API_KEY:
                st.error("❌ Missing LLMWhisperer API key. Please set it in Streamlit secrets.")
            else:
                result = extract_llm_pdf(uploaded_pdf, LLM_API_KEY)
                st.success("✅ LLMWhisperer processing complete.")
                st.subheader("LLMWhisperer Extracted Output:")
                st.json(result)

        except Exception as e:
            st.error(f"❌ Error processing LLMWhisperer extraction: {str(e)}")
